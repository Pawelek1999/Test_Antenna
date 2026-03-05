from fastapi import FastAPI, BackgroundTasks, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, StreamingResponse
from pydantic import BaseModel
from typing import List, Dict, Any
import time
import json
import os
import io
from pathlib import Path
import openpyxl
import re
from testowy import run_antenna_test  # Import funkcji testowej z testowy.py
from test_state import TestState, RESULT_FILE
from paths import CONFIG


app = FastAPI()

# --- Konfiguracja CORS ---
# Pozwala frontendowi (React) na komunikację z backendem
origins = [
    "http://localhost:5173",
    "http://localhost:3000",
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

REPORT_TEMPLATE_PATH = CONFIG / "report_template.xlsx"
UPLOADED_TEMPLATE_PATH = CONFIG / "uploaded_template.xlsx"
FREQUENCY_JSON_PATH = CONFIG / "frequency.json"
RUNTIME_PARAMS_PATH = CONFIG / "runtime_params.json"
HARDWARE_CONFIG_PATH = CONFIG / "hardware_config.json"
TEST_CONFIG_PATH = CONFIG / "test_config.json"
WE_CONFIG_PATH = CONFIG / "we_config.json"
test_state = TestState()

@app.post("/start-test")
async def start_test(background_tasks: BackgroundTasks):
    if test_state.is_active():
        return JSONResponse(status_code=409, content={"message": "Test jest już w toku."})

    # Usuń stary plik wyników, jeśli istnieje
    if os.path.exists(RESULT_FILE):
        os.remove(RESULT_FILE)
    
    test_state.start()
    background_tasks.add_task(run_antenna_test, test_state)
    return {"message": "Test uruchomiony w tle"}

@app.post("/stop-test")
async def stop_test():
    """Zatrzymuje aktualnie działający test."""
    if not test_state.is_active():
        return {"message": "Żaden test nie jest aktualnie uruchomiony."}
    
    test_state.stop()
    return {"message": "Wysłano sygnał zatrzymania testu."}

@app.get("/check-status")
async def check_status():
    return {"is_running": test_state.is_active(), "results_ready": os.path.exists(RESULT_FILE)}

@app.get("/download-data")
async def download_data():
    if not os.path.exists(RESULT_FILE):
        return JSONResponse(status_code=404, content={"message": "Brak wyników"})
    
    with open(RESULT_FILE, "r") as f:
        data = json.load(f)
    return data


@app.get("/frequencies")
async def get_frequencies():
    """Zwraca aktualną konfigurację częstotliwości z pliku w backendzie."""
    if not os.path.exists(FREQUENCY_JSON_PATH):
        return []
    try:
        with open(FREQUENCY_JSON_PATH, "r", encoding="utf-8") as f:
            content = f.read()
            return json.loads(content) if content.strip() else []
    except json.JSONDecodeError:
        return []
 
def _fill_workbook_with_results(workbook: openpyxl.Workbook, test_results: list):
    """Helper function to populate an Excel workbook with test results."""
    sheet = workbook.worksheets[0]  # Select the first sheet

    # --- Populate data ---
    start_row = 22
    for index, row_data in enumerate(test_results):
        current_row = start_row + index
        # Column mapping: E=5, F=6, G=7, H=8
        sheet.cell(row=current_row, column=5).value = row_data.get("genPolarH_act")
        sheet.cell(row=current_row, column=6).value = row_data.get("genPolarH_stop")
        sheet.cell(row=current_row, column=7).value = row_data.get("genPolarV_act")
        sheet.cell(row=current_row, column=8).value = row_data.get("genPolarV_stop")


def _generate_excel_response(workbook: openpyxl.Workbook) -> StreamingResponse:
    """Saves a workbook to a memory buffer and returns it as a StreamingResponse."""
    output = io.BytesIO()
    workbook.save(output)
    # --- Weryfikacja ---
    # Sprawdzamy i drukujemy w konsoli backendu rozmiar bufora po zapisie.
    print(f"Plik Excel wygenerowany w pamięci. Rozmiar bufora: {output.tell()} bajtów.")
    output.seek(0)

    filename = f"Raport_Anteny_{int(time.time())}.xlsx"
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )


@app.post("/upload-template")
async def upload_template(file: UploadFile = File(...)):
    """
    1. Zapisuje przesłany plik jako szablon w pamięci backendu (na dysku).
    2. Aktualizuje frequency.json na podstawie danych z kolumn T-Y tego pliku.
    """
    try:
        content = await file.read()
        
        # Upewnij się, że katalog docelowy istnieje
        UPLOADED_TEMPLATE_PATH.parent.mkdir(parents=True, exist_ok=True)

        # 1. Zapisz plik na serwerze (jako szablon do późniejszego użycia)
        with open(UPLOADED_TEMPLATE_PATH, "wb") as f:
            f.write(content)

        # 2. Przetwórz częstotliwości (logika parsowania kolumn T-Y)
        workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        sheet = workbook.active

        # Helper do obsługi scalonych komórek
        def get_value(r, c):
            cell = sheet.cell(row=r, column=c)
            if cell.value is not None:
                return cell.value
            # Sprawdź czy komórka jest częścią scalonego obszaru
            for merged_range in sheet.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    # Zwróć wartość z lewego górnego rogu scalenia
                    return sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value
            return None

        # Pobierz klucze z kolumny T (indeks 20) w wierszach 3-8
        keys = []
        for r in range(3, 9):
            val = get_value(r, 20) # Kolumna T (klucze)
            if val:
                # Normalizacja klucza: znajduje wszystkie "słowa" (ciągi liter i cyfr),
                # konwertuje na małe litery i łączy je podkreśleniem.
                # Przykład: "Frequency (MHz)" -> "frequency_mhz"
                words = re.findall(r'\w+', str(val).lower())
                key = "_".join(words)
                keys.append(key)
            else:
                keys.append(f"field_{r}")

        frequencies_data = []
        current_id = 1
        
        # Iteracja po kolumnach od V (22) do AD (30) - każda kolumna to obiekt
        for col in range(22, 31):
            # Sprawdź czy obiekt istnieje (czy ma wartość w pierwszym wierszu kluczy - wiersz 3)
            if get_value(3, col) is None:
                continue

            item = {"id": current_id}
            for i, r in enumerate(range(3, 9)):
                key = keys[i]
                val = get_value(r, col)
                item[key] = val

            frequencies_data.append(item)
            current_id += 1

        # Zapisz nowe dane do pliku JSON
        with open(FREQUENCY_JSON_PATH, "w", encoding="utf-8") as f:
            json.dump(frequencies_data, f, indent=2, ensure_ascii=False)

        return {"message": f"Szablon zapisany. Zaktualizowano {len(frequencies_data)} częstotliwości."}

    except HTTPException as e:
        raise e # Przekaż dalej wyjątki HTTP z logiką walidacji
    except Exception as e:
        print(f"Błąd przetwarzania szablonu Excel: {e}")
        raise HTTPException(status_code=500, detail=f"Wystąpił błąd serwera podczas przetwarzania pliku: {str(e)}")


@app.get("/generate-report")
async def generate_report():
    """
    Generuje raport Excel, używając wcześniej przesłanego szablonu (upload-template)
    i wypełniając go aktualnymi wynikami testu.
    """
    if not os.path.exists(RESULT_FILE):
        raise HTTPException(status_code=404, detail="Brak wyników testu do wyeksportowania.")

    if not os.path.exists(UPLOADED_TEMPLATE_PATH):
        raise HTTPException(status_code=400, detail="Nie załadowano szablonu Excel. Przeciągnij plik w sekcji eksportu.")

    with open(RESULT_FILE, "r") as f:
        test_results = json.load(f)

    try:
        workbook = openpyxl.load_workbook(UPLOADED_TEMPLATE_PATH)
        _fill_workbook_with_results(workbook, test_results)
        return _generate_excel_response(workbook)

    except Exception as e:
        print(f"Błąd generowania raportu Excel: {e}")
        raise HTTPException(status_code=500, detail=f"Błąd serwera: {str(e)}")


# --- Obsługa parametrów Runtime ---

class RuntimeParams(BaseModel):
    start_power_dbm: float
    end_power_dbm: float
    power_step_db: float
    start_table_position: int
    start_malt_height: int
    silent_search_reduction_db: float

@app.post("/save-runtime-params")
async def save_runtime_params(params: RuntimeParams):
    """Zapisuje parametry wejściowe testu do pliku JSON w folderze config."""
    try:
        with open(RUNTIME_PARAMS_PATH, "w", encoding="utf-8") as f:
            json.dump(params.dict(), f, indent=2)
        return {"message": "Parametry runtime zostały zapisane."}
    except Exception as e:
        print(f"Błąd zapisu runtime_params: {e}")
        raise HTTPException(status_code=500, detail=f"Błąd zapisu pliku: {str(e)}")


# --- Obsługa konfiguracji sprzętowej ---

class HardwareConfig(BaseModel):
    analog_channel: int
    voltage_threshold_v: float
    safe_stop_power_dbm: float

@app.post("/save-hardware-config")
async def save_hardware_config(config: HardwareConfig):
    """Zapisuje konfigurację sprzętową do pliku JSON w folderze config."""
    try:
        with open(HARDWARE_CONFIG_PATH, "w", encoding="utf-8") as f:
            json.dump(config.dict(), f, indent=2)
        return {"message": "Konfiguracja sprzętowa została zapisana."}
    except Exception as e:
        print(f"Błąd zapisu hardware_config: {e}")
        raise HTTPException(status_code=500, detail=f"Błąd zapisu pliku: {str(e)}")


# --- Obsługa konfiguracji testu (Multi-Sheet) ---

@app.post("/save-test-config")
async def save_test_config(config: List[Dict[str, Any]]):
    """Zapisuje wygenerowaną konfigurację testu (lista obiektów) do pliku JSON."""
    try:
        with open(TEST_CONFIG_PATH, "w", encoding="utf-8") as f:
            json.dump(config, f, indent=2)
        return {"message": f"Zapisano konfigurację testu ({len(config)} arkuszy)."}
    except Exception as e:
        print(f"Błąd zapisu test_config: {e}")
        raise HTTPException(status_code=500, detail=f"Błąd zapisu pliku: {str(e)}")


# --- Obsługa scalania konfiguracji (WE Config) ---

@app.get("/get-hardware-config")
async def get_hardware_config():
    if not os.path.exists(HARDWARE_CONFIG_PATH):
        return {}
    try:
        with open(HARDWARE_CONFIG_PATH, "r", encoding="utf-8") as f:
            content = f.read()
            return json.loads(content) if content.strip() else {}
    except json.JSONDecodeError:
        return {}

@app.get("/get-runtime-params")
async def get_runtime_params():
    if not os.path.exists(RUNTIME_PARAMS_PATH):
        return {}
    try:
        with open(RUNTIME_PARAMS_PATH, "r", encoding="utf-8") as f:
            content = f.read()
            return json.loads(content) if content.strip() else {}
    except json.JSONDecodeError:
        return {}

@app.get("/get-test-config")
async def get_test_config():
    if not os.path.exists(TEST_CONFIG_PATH):
        return []
    try:
        with open(TEST_CONFIG_PATH, "r", encoding="utf-8") as f:
            content = f.read()
            return json.loads(content) if content.strip() else []
    except json.JSONDecodeError:
        return []

@app.post("/save-we-config")
async def save_we_config(config: List[Dict[str, Any]]):
    """
    Zapisuje scalony plik konfiguracyjny (we_config.json).
    Oczekuje listy obiektów, gdzie każdy obiekt zawiera kompletne dane dla jednego arkusza.
    """
    try:
        # Zapisz do pliku we_config.json
        with open(WE_CONFIG_PATH, "w", encoding="utf-8") as f:
            json.dump(config, f, indent=2)
        
        return {"message": "Plik we_config.json został pomyślnie wygenerowany i zapisany."}
    except Exception as e:
        print(f"Błąd zapisu we_config: {e}")
        raise HTTPException(status_code=500, detail=f"Błąd zapisu pliku: {str(e)}")